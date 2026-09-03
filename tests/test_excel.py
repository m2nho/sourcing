from openpyxl import load_workbook

from sourcing.excel import EXCEL_HEADERS, write_xlsx
from sourcing.store import (
    SOURCE_MAP_PHONE_GUESS,
    SOURCE_SITE_CONFIRMS_MAP,
    SOURCE_SITE_LINK,
    PlaceRecord,
)


def rec(name, status, source, phone="+6281100000001", address="Jl. Contoh 1", wa="",
        cid=None, website="", maps_url=""):
    return PlaceRecord(
        place_cid=cid or f"0xa:{name}",
        name=name,
        address=address,
        phone_e164=phone,
        whatsapp_status=status,
        source=source,
        wa_link=wa,
        website=website,
        maps_url=maps_url,
    )


def sheet(path):
    return load_workbook(path).active


def test_headers_are_the_slim_set(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx([rec("Klinik A", "confirmed", SOURCE_SITE_LINK)], out)
    header = [c.value for c in sheet(out)[1]]
    assert header == EXCEL_HEADERS
    assert header == [
        "병원명", "위치", "전화번호", "WhatsApp 링크", "상태", "근거",
        "WhatsApp 프로필", "검색어", "홈페이지", "구글맵",
    ]


def test_source_is_written_as_a_human_label(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx(
        [
            rec("A", "confirmed", SOURCE_SITE_CONFIRMS_MAP),
            rec("B", "confirmed", SOURCE_SITE_LINK),
            rec("C", "candidate", SOURCE_MAP_PHONE_GUESS),
        ],
        out,
    )
    labels = [row[5].value for row in sheet(out).iter_rows(min_row=2)]
    assert labels == ["홈페이지+맵 일치", "홈페이지 링크", "맵 번호 추정"]


def test_rows_carry_both_links(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx(
        [rec("A", "confirmed", SOURCE_SITE_LINK,
             website="https://klinik.co.id/", maps_url="https://maps.google.com/x")],
        out,
    )
    row = [c.value for c in sheet(out)[2]]
    assert row[8] == "https://klinik.co.id/"
    assert row[9] == "https://maps.google.com/x"


def test_missing_links_leave_the_cell_empty(tmp_path):
    # openpyxl은 빈 문자열을 빈 셀로 저장하고 None으로 읽는다. 사용자에게는
    # 빈 칸으로 보이므로 정상이다 - 'None' 같은 글자가 새지 않는 것이 요건이다.
    out = tmp_path / "leads.xlsx"
    write_xlsx([rec("A", "confirmed", SOURCE_SITE_LINK)], out)
    row = [c.value for c in sheet(out)[2]]
    assert not row[8]
    assert not row[9]


def test_rows_carry_name_and_location(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx([rec("Klinik Contoh", "confirmed", SOURCE_SITE_LINK, address="Jl. Sudirman 5")], out)
    row = [c.value for c in sheet(out)[2]]
    assert row[0] == "Klinik Contoh"
    assert row[1] == "Jl. Sudirman 5"


def test_confirmed_rows_come_first(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx(
        [
            rec("추정", "candidate", SOURCE_MAP_PHONE_GUESS),
            rec("확정", "confirmed", SOURCE_SITE_LINK),
        ],
        out,
    )
    names = [row[0].value for row in sheet(out).iter_rows(min_row=2)]
    assert names == ["확정", "추정"]


def test_unlikely_rows_are_left_out(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx(
        [rec("연락가능", "confirmed", SOURCE_SITE_LINK), rec("근거없음", "unlikely", "")], out
    )
    names = [row[0].value for row in sheet(out).iter_rows(min_row=2)]
    assert names == ["연락가능"]


def test_empty_input_still_writes_a_header(tmp_path):
    out = tmp_path / "leads.xlsx"
    assert write_xlsx([], out) == 0
    assert [c.value for c in sheet(out)[1]] == EXCEL_HEADERS


def test_returns_written_row_count(tmp_path):
    out = tmp_path / "leads.xlsx"
    written = write_xlsx(
        [rec("A", "confirmed", SOURCE_SITE_LINK), rec("B", "candidate", SOURCE_MAP_PHONE_GUESS)],
        out,
    )
    assert written == 2


def test_write_from_jsonl_rebuilds_the_sheet(tmp_path):
    """취소된 작업이나 예전 데이터도 JSONL만 있으면 엑셀을 다시 뽑을 수 있어야 한다."""
    import json

    from sourcing.excel import write_xlsx_from_jsonl

    raw = tmp_path / "run.raw.jsonl"
    with raw.open("w", encoding="utf-8") as fh:
        for r in (
            rec("확정", "confirmed", SOURCE_SITE_LINK),
            rec("추정", "candidate", SOURCE_MAP_PHONE_GUESS),
            rec("제외", "unlikely", ""),
        ):  # place_cid가 서로 달라야 JsonlStore의 중복 제거에 걸리지 않는다
            fh.write(json.dumps(r.__dict__, ensure_ascii=False) + "\n")

    out = tmp_path / "run.xlsx"
    assert write_xlsx_from_jsonl(raw, out) == 2
    names = [row[0].value for row in sheet(out).iter_rows(min_row=2)]
    assert names == ["확정", "추정"]


def test_write_from_jsonl_dedupes_like_the_store(tmp_path):
    import json

    from sourcing.excel import write_xlsx_from_jsonl

    raw = tmp_path / "run.raw.jsonl"
    duplicate = rec("A", "confirmed", SOURCE_SITE_LINK, cid="0xa:0xb")
    with raw.open("w", encoding="utf-8") as fh:
        fh.write(json.dumps(duplicate.__dict__, ensure_ascii=False) + "\n")
        fh.write(json.dumps(duplicate.__dict__, ensure_ascii=False) + "\n")
    assert write_xlsx_from_jsonl(raw, tmp_path / "run.xlsx") == 1


def test_write_from_jsonl_on_missing_file_is_zero(tmp_path):
    from sourcing.excel import write_xlsx_from_jsonl

    out = tmp_path / "run.xlsx"
    assert write_xlsx_from_jsonl(tmp_path / "nope.jsonl", out) == 0
    assert [c.value for c in sheet(out)[1]] == EXCEL_HEADERS


def test_each_grade_gets_its_own_colour(tmp_path):
    """링크는 모두에게 주므로, 신뢰도 차이는 색으로 한눈에 보여야 한다."""
    from sourcing.store import SOURCE_PROFILE

    out = tmp_path / "leads.xlsx"
    write_xlsx(
        [
            rec("확정", "confirmed", SOURCE_SITE_LINK),
            rec("검증", "verified", SOURCE_PROFILE),
            rec("추정", "candidate", SOURCE_MAP_PHONE_GUESS),
        ],
        out,
    )
    ws = sheet(out)
    colours = {row[4].value: row[4].fill.start_color.rgb for row in ws.iter_rows(min_row=2)}
    assert len({c for c in colours.values()}) == 3, "세 등급의 색이 서로 달라야 한다"


def test_every_row_carries_a_clickable_link(tmp_path):
    from sourcing.store import SOURCE_PROFILE

    out = tmp_path / "leads.xlsx"
    write_xlsx(
        [
            rec("확정", "confirmed", SOURCE_SITE_LINK, wa="https://wa.me/6281100000001"),
            rec("추정", "candidate", SOURCE_MAP_PHONE_GUESS, wa="https://wa.me/6281100000002"),
        ],
        out,
    )
    links = [row[3].value for row in sheet(out).iter_rows(min_row=2)]
    assert all(l and l.startswith("https://wa.me/") for l in links)


def test_workbook_carries_a_legend_sheet(tmp_path):
    """엑셀만 받은 사람이 상태와 근거의 뜻을 물어보지 않아도 알 수 있어야 한다."""
    from sourcing.excel import LEGEND_SHEET
    from sourcing.store import SOURCE_LABELS

    out = tmp_path / "leads.xlsx"
    write_xlsx([rec("A", "confirmed", SOURCE_SITE_LINK)], out)
    book = load_workbook(out)
    assert LEGEND_SHEET in book.sheetnames

    text = "\n".join(
        str(c.value) for row in book[LEGEND_SHEET].iter_rows() for c in row if c.value
    )
    for label in SOURCE_LABELS.values():
        if label != "근거 없음":  # unlikely는 엑셀에 나오지 않는다
            assert label in text, f"근거 '{label}' 설명이 범례에 없다"
    for grade in ("확정", "검증", "추정"):
        assert grade in text


def test_legend_explains_the_colours(tmp_path):
    from sourcing.excel import LEGEND_SHEET, _STATUS_FILLS

    out = tmp_path / "leads.xlsx"
    write_xlsx([rec("A", "confirmed", SOURCE_SITE_LINK)], out)
    ws = load_workbook(out)[LEGEND_SHEET]
    used = {c.fill.start_color.rgb for row in ws.iter_rows() for c in row}
    for fill in _STATUS_FILLS.values():
        assert f"00{fill.start_color.rgb[-6:]}" in used or fill.start_color.rgb in used


def test_data_sheet_is_the_first_one(tmp_path):
    out = tmp_path / "leads.xlsx"
    write_xlsx([rec("A", "confirmed", SOURCE_SITE_LINK)], out)
    book = load_workbook(out)
    assert book.sheetnames[0] == "리드"


def test_search_term_column_shows_where_the_lead_came_from(tmp_path):
    """여러 지역을 합친 파일에서는 어느 검색으로 찾았는지가 중요하다."""
    out = tmp_path / "leads.xlsx"
    r = rec("A", "confirmed", SOURCE_SITE_LINK)
    r.query = "aesthetic clinic Knightsbridge"
    write_xlsx([r], out)
    ws = sheet(out)
    assert "검색어" in [c.value for c in ws[1]]
    idx = [c.value for c in ws[1]].index("검색어")
    assert ws.cell(2, idx + 1).value == "aesthetic clinic Knightsbridge"
