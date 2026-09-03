"""영업 담당자가 바로 여는 엑셀 파일을 만든다.

CSV는 전체 레코드를 보존하는 원본이고, 엑셀은 실제로 연락할 목록이다.
그래서 컬럼이 다르다 — 병원명과 위치, 연락 수단, 그리고 그 번호를 어디서
얻었는지(근거)만 남긴다.

근거를 남기는 이유: confirmed 안에도 성격이 다른 것들이 섞여 있다. 홈페이지가
맵 번호를 확인해준 것과, 홈페이지에만 있던 번호와, 맵 링크에서 나온 것은
신뢰도가 다르다. 담당자가 어디부터 걸지 정할 수 있어야 한다.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sourcing.store import (
    SOURCE_MAP_LINK,
    SOURCE_MAP_PHONE_GUESS,
    SOURCE_PROFILE,
    SOURCE_PROFILE_MISMATCH,
    SOURCE_SITE_CONFIRMS_MAP,
    SOURCE_SITE_LINK,
    SOURCE_LABELS,
    PROFILE_LABELS,
    JsonlStore,
    PlaceRecord,
)

EXCEL_HEADERS = [
    "병원명", "위치", "전화번호", "WhatsApp 링크", "상태", "근거",
    "WhatsApp 프로필", "검색어", "홈페이지", "구글맵",
]

#: 연락할 수 없는 곳은 목록에 넣지 않는다. 원본은 CSV에 그대로 남아 있다.
EXPORTED_STATUSES = ("confirmed", "verified", "candidate")

#: 확정을 먼저 보여준다. 담당자가 위에서부터 걸면 된다.
_STATUS_ORDER = {"confirmed": 0, "verified": 1, "candidate": 2}

_STATUS_LABELS = {"confirmed": "확정", "verified": "검증", "candidate": "추정"}

_COLUMN_WIDTHS = (38, 52, 18, 30, 8, 16, 32, 30, 42, 42)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")

#: 등급별 색. 링크는 모두에게 주므로 신뢰도 차이는 색으로 알린다.
_STATUS_FILLS = {
    "confirmed": PatternFill("solid", fgColor="C6E0B4"),  # 초록 - 업체가 선언
    "verified": PatternFill("solid", fgColor="BDD7EE"),   # 파랑 - 프로필 확인
    "candidate": PatternFill("solid", fgColor="FFE699"),  # 노랑 - 미확인 추측
}


LEGEND_SHEET = "범례"

#: 등급이 무엇을 뜻하는지. 측정 결과가 아니라 정의만 적는다 - 수치는
#: 실행마다 달라지지만 정의는 그대로다.
_GRADE_NOTES = [
    ("확정", "confirmed", "업체가 자기 웹사이트에 WhatsApp 링크를 공개해 둔 번호다"),
    ("검증", "verified", "웹사이트에 공개돼 있진 않지만, 번호를 조회하니 WhatsApp 프로필 이름이 상호와 일치하는 번호다"),
    ("추정", "candidate", "구글맵 대표번호가 모바일 번호대라 WhatsApp일 가능성이 있는 번호다. 확인되지는 않았다"),
]

#: 근거 설명. 같은 등급 안에서도 신뢰도가 갈린다.
_SOURCE_NOTES = [
    (SOURCE_SITE_CONFIRMS_MAP, "홈페이지의 wa.me 번호가 구글맵 대표번호와 같다. 두 출처가 서로를 확인한 것으로 가장 강하다"),
    (SOURCE_SITE_LINK, "홈페이지에서 찾은 선언(wa.me 링크 또는 WhatsApp이라 표시된 tel: 링크). 맵에는 없던 번호다"),
    (SOURCE_MAP_LINK, "구글맵의 웹사이트 칸 자체가 wa.me 링크였다. 드물다"),
    (SOURCE_PROFILE, "번호를 wa.me에서 조회하니 프로필 이름이 상호와 일치했다"),
    (SOURCE_PROFILE_MISMATCH, "번호는 WhatsApp에 있지만 프로필 이름이 상호와 다르다. 원장 개인 이름일 수 있으니 링크를 눌러 확인하라"),
    (SOURCE_MAP_PHONE_GUESS, "구글맵 대표번호가 모바일 번호대다. 실제 등록 여부는 확인되지 않았다"),
]

#: WhatsApp 프로필 칸에 이름 대신 들어가는 값의 뜻.
_PROFILE_NOTES = [
    ("(개인 또는 미등록)", "번호를 조회했지만 이름이 보이지 않았다. WhatsApp은 비즈니스 계정의 이름만 공개한다 - 개인 계정으로 쓰는 곳이거나 미등록이며, 이 둘은 구별할 수 없다"),
    ("(확인 실패)", "조회 자체가 되지 않았다(접속 실패·시간 초과). 다시 조회하면 결과가 나올 수 있다"),
]

_NOTE_HEADER = Font(bold=True, color="FFFFFF")


def _write_legend(workbook) -> None:
    """상태와 근거의 뜻을 적은 시트. 엑셀만 받은 사람도 읽고 판단할 수 있게 한다."""
    sheet = workbook.create_sheet(LEGEND_SHEET)

    sheet.append(["상태", "뜻"])
    for cell in sheet[1]:
        cell.font = _NOTE_HEADER
        cell.fill = _HEADER_FILL
    for label, status, meaning in _GRADE_NOTES:
        sheet.append([label, meaning])
        sheet.cell(sheet.max_row, 1).fill = _STATUS_FILLS[status]

    sheet.append([])
    sheet.append(["근거", "뜻"])
    for cell in sheet[sheet.max_row]:
        cell.font = _NOTE_HEADER
        cell.fill = _HEADER_FILL
    for source, meaning in _SOURCE_NOTES:
        sheet.append([SOURCE_LABELS[source], meaning])

    sheet.append([])
    sheet.append(["WhatsApp 프로필", "뜻"])
    for cell in sheet[sheet.max_row]:
        cell.font = _NOTE_HEADER
        cell.fill = _HEADER_FILL
    for label, meaning in _PROFILE_NOTES:
        sheet.append([label, meaning])

    sheet.append([])
    sheet.append(["참고", "모든 행에 클릭 가능한 wa.me 링크가 있다. 신뢰도 차이는 상태 색과 근거로 표시한다"])
    sheet.append(["", "연락이 불가능한 곳(번호 없음·미등록 유선)은 이 파일에 넣지 않는다. 전체는 CSV에 있다"])

    for column, width in (("A", 20), ("B", 96)):
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def _status_of(label: str | None) -> str:
    """엑셀에 찍힌 한국어 라벨을 내부 상태값으로 되돌린다."""
    for status, text in _STATUS_LABELS.items():
        if text == label:
            return status
    return ""


def write_xlsx(records: Iterable[PlaceRecord], path: Path) -> int:
    """연락 가능한 레코드를 엑셀로 쓴다. 쓴 행 수를 돌려준다."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    rows = sorted(
        (r for r in records if r.whatsapp_status in EXPORTED_STATUSES),
        key=lambda r: (_STATUS_ORDER.get(r.whatsapp_status, 9), r.name),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "리드"
    sheet.append(EXCEL_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL

    for record in rows:
        sheet.append(
            [
                record.name,
                record.address,
                record.phone_e164,
                record.wa_link,
                _STATUS_LABELS.get(record.whatsapp_status, record.whatsapp_status),
                SOURCE_LABELS.get(record.source, record.source),
                record.profile_name or PROFILE_LABELS.get(record.profile_checked, ""),
                record.query,
                record.website,
                record.maps_url,
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        fill = _STATUS_FILLS.get(_status_of(row[4].value))
        if fill is not None:
            row[4].fill = fill

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}{sheet.max_row}"

    _write_legend(workbook)
    workbook.active = 0  # 열었을 때 리드 시트가 먼저 보이게
    workbook.save(path)
    return len(rows)


def write_xlsx_from_jsonl(raw_jsonl: Path, out_path: Path) -> int:
    """재개 원장(JSONL)에서 엑셀을 다시 만든다. 쓴 행 수를 돌려준다.

    수집이 취소되면 CLI의 마무리 코드에 도달하지 못해 엑셀이 남지 않는다.
    JSONL은 레코드마다 flush되므로 그것만 있으면 언제든 복원할 수 있다.
    """
    return write_xlsx(JsonlStore(Path(raw_jsonl)).records(), out_path)
